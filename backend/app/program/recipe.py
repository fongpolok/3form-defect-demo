"""
Port of the old app's Excel/DataGridView "recipe" grid: an ordered list of
inspection steps, each a robot pose plus the camera/light/focus/detector
settings to use there. `Form1.cs` stored exactly these fields per row
(Action, Exposure, Brightness, Contrast, Joint 1-6, Pattern Width/Rotation/
Shift/Intensity, Focus Position, Img Path, vrws Path) and replayed them via
a "Do" button — see `legacy-csharp/README.md`. Here it's JSON-native (so the
frontend can edit it directly) with .xlsx import/export via openpyxl for
operators used to the old spreadsheet-shaped workflow.
"""
from __future__ import annotations

import json
import math
import uuid
from pathlib import Path

import openpyxl
from pydantic import BaseModel, Field

from app.config import BACKEND_DIR
from app.logging_setup import get_logger

logger = get_logger(__name__)

RECIPES_DIR = BACKEND_DIR / "data" / "recipes"

XLSX_COLUMNS = [
    "Name", "Stay Only",
    "Joint 1 (deg)", "Joint 2 (deg)", "Joint 3 (deg)", "Joint 4 (deg)", "Joint 5 (deg)", "Joint 6 (deg)",
    "Speed", "Acceleration",
    "Exposure", "Brightness", "Contrast", "Focus Position",
    "Pattern Width", "Pattern Rotation", "Pattern Shift", "Pattern Intensity",
    "Detector",
]


class LightPatternSettings(BaseModel):
    width: int = 30
    rotation: int = 0
    shift: int = 0
    intensity: int = 255


class RecipeStep(BaseModel):
    id: str = Field(default_factory=lambda: uuid.uuid4().hex[:8])
    name: str = "Step"
    # "Stay" in the old grid meant: move here, but don't capture/inspect —
    # useful for approach/retract waypoints.
    stay_only: bool = False

    joint_positions_deg: list[float] = Field(default_factory=lambda: [0.0] * 6)
    speed: float = 0.5
    acceleration: float = 0.5

    camera_exposure: float | None = None
    camera_brightness: float | None = None
    camera_contrast: float | None = None
    focus_position: int | None = None

    light_pattern: LightPatternSettings = Field(default_factory=LightPatternSettings)

    # Which DefectDetector to run at this step: "classical" | "patchcore" | "yolo" | None
    detector: str | None = None


class Recipe(BaseModel):
    name: str = "default"
    steps: list[RecipeStep] = Field(default_factory=list)


class RecipeStore:
    """Simple file-backed store — one JSON file per recipe under data/recipes/."""

    def __init__(self, directory: Path = RECIPES_DIR) -> None:
        self.directory = directory
        self.directory.mkdir(parents=True, exist_ok=True)

    def _path(self, name: str) -> Path:
        safe_name = "".join(c for c in name if c.isalnum() or c in "-_") or "default"
        return self.directory / f"{safe_name}.json"

    def list_names(self) -> list[str]:
        return sorted(p.stem for p in self.directory.glob("*.json"))

    def load(self, name: str) -> Recipe:
        path = self._path(name)
        if not path.exists():
            logger.info("No saved recipe %r yet — returning an empty one", name)
            return Recipe(name=name)
        data = json.loads(path.read_text(encoding="utf-8"))
        return Recipe(**data)

    def save(self, recipe: Recipe) -> None:
        path = self._path(recipe.name)
        path.write_text(recipe.model_dump_json(indent=2), encoding="utf-8")
        logger.info("Saved recipe %r (%d steps) to %s", recipe.name, len(recipe.steps), path)

    def delete(self, name: str) -> bool:
        path = self._path(name)
        if path.exists():
            path.unlink()
            logger.info("Deleted recipe %r", name)
            return True
        return False


def export_xlsx(recipe: Recipe, path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = recipe.name[:31] or "recipe"
    ws.append(XLSX_COLUMNS)
    for step in recipe.steps:
        ws.append([
            step.name, step.stay_only,
            *step.joint_positions_deg,
            step.speed, step.acceleration,
            step.camera_exposure, step.camera_brightness, step.camera_contrast, step.focus_position,
            step.light_pattern.width, step.light_pattern.rotation, step.light_pattern.shift, step.light_pattern.intensity,
            step.detector,
        ])
    wb.save(path)
    logger.info("Exported recipe %r to %s", recipe.name, path)


def import_xlsx(path: Path, name: str | None = None) -> Recipe:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return Recipe(name=name or path.stem)

    header = [str(h) for h in rows[0]]
    steps: list[RecipeStep] = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        values = dict(zip(header, row))
        steps.append(RecipeStep(
            name=str(values.get("Name") or "Step"),
            stay_only=bool(values.get("Stay Only") or False),
            joint_positions_deg=[
                _to_float(values.get(f"Joint {i} (deg)")) for i in range(1, 7)
            ],
            speed=_to_float(values.get("Speed")) or 0.5,
            acceleration=_to_float(values.get("Acceleration")) or 0.5,
            camera_exposure=_to_float(values.get("Exposure")),
            camera_brightness=_to_float(values.get("Brightness")),
            camera_contrast=_to_float(values.get("Contrast")),
            focus_position=_to_int(values.get("Focus Position")),
            light_pattern=LightPatternSettings(
                width=_to_int(values.get("Pattern Width")) or 30,
                rotation=_to_int(values.get("Pattern Rotation")) or 0,
                shift=_to_int(values.get("Pattern Shift")) or 0,
                intensity=_to_int(values.get("Pattern Intensity")) or 255,
            ),
            detector=values.get("Detector") or None,
        ))
    return Recipe(name=name or path.stem, steps=steps)


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    f = float(value)
    return None if math.isnan(f) else f


def _to_int(value) -> int | None:
    f = _to_float(value)
    return None if f is None else int(f)
